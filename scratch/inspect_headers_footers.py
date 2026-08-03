import openpyxl
import os

folder = r"c:\Users\JULIAN\Desktop\GravyLocal2.0\DatosReferencia\reportes"
files = [
    "Estado de resultados.xlsx",
    "estado de situacion financiera.xlsx",
    "flujo de caja.xlsx",
    "notas a estados financieros.xlsx"
]

for filename in files:
    path = os.path.join(folder, filename)
    print(f"\n==========================================")
    print(f"FILE: {filename}")
    print(f"==========================================")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    print("Header rows (1 to 10):")
    for r in range(1, min(11, max_row + 1)):
        row_vals = [sheet.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v is not None for v in row_vals):
            print(f"  Row {r:02d}: {row_vals[:8]} ... (total cols: {max_col})")
            
    print("\nFooter rows (last 15):")
    for r in range(max(1, max_row - 14), max_row + 1):
        row_vals = [sheet.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v is not None for v in row_vals):
            print(f"  Row {r:02d}: {row_vals[:8]} ... (total cols: {max_col})")
            
    # Look for signature lines
    print("\nSignature-like cells:")
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = sheet.cell(r, c).value
            if val and any(keyword in str(val).lower() for keyword in ["firma", "representante", "contador", "revisor", "t.p", "tp.", "matrícula"]):
                print(f"  Cell({r},{c}) [{val}]: {sheet.cell(r, c).alignment.horizontal}")
