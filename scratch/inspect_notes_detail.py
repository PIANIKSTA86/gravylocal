import openpyxl
import os

path = r"c:\Users\JULIAN\Desktop\GravyLocal2.0\DatosReferencia\reportes\notas a estados financieros.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

print(f"Sheet dimensions: {sheet.dimensions}")

# Print rows that have headers or seem to define notes
print("\n--- DETECTED SECTIONS / NOTES ---")
for r in range(1, max_row + 1):
    val = sheet.cell(r, 2).value # Usually notes are in column B (Col 2)
    if val:
        val_str = str(val).strip()
        # Look for headers like "1. ", "2. ", "Nota", "NOTAS", or bold text with numbers
        is_note_header = False
        if val_str and (val_str[0].isdigit() and ('.' in val_str[:3] or ')' in val_str[:3])):
            is_note_header = True
        elif "nota" in val_str.lower():
            is_note_header = True
        elif val_str.isupper() and len(val_str) > 3 and sheet.cell(r, 2).font.bold:
            is_note_header = True
            
        if is_note_header:
            print(f"Row {r:03d} (Col B): {val_str}")
            
# Let's inspect some of the tables inside this sheet.
# For example, look at rows around some sections. Let's list rows 15 to 45 to see how tables look.
print("\n--- SAMPLE ROWS 15 TO 45 ---")
for r in range(15, 46):
    row_vals = [sheet.cell(r, c).value for c in range(1, max_col + 1)]
    if any(v is not None for v in row_vals):
        # Only print non-empty columns
        non_empty = [(c+1, v) for c, v in enumerate(row_vals) if v is not None]
        print(f"Row {r:03d}: {non_empty[:6]}")
        
# Inspect rows around row 80 to 110 (which has some tables)
print("\n--- SAMPLE ROWS 80 TO 110 ---")
for r in range(80, 111):
    row_vals = [sheet.cell(r, c).value for c in range(1, max_col + 1)]
    if any(v is not None for v in row_vals):
        non_empty = [(c+1, v) for c, v in enumerate(row_vals) if v is not None]
        print(f"Row {r:03d}: {non_empty[:6]}")
