import openpyxl
import os

folder = r"c:\Users\JULIAN\Desktop\GravyLocal2.0\DatosReferencia\reportes"
files = [
    "Estado de resultados.xlsx",
    "estado de situacion financiera.xlsx",
    "flujo de caja.xlsx",
    "notas a estados financieros.xlsx"
]

for filename in files:
    path = os.path.join(folder, filename)
    print(f"\n==========================================")
    print(f"FILE: {filename}")
    print(f"==========================================")
    
    if not os.path.exists(path):
        print(f"File not found: {path}")
        continue
        
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheets:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name} (Dimensions: {sheet.dimensions})")
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        print(f"Rows: {max_row}, Columns: {max_col}")
        
        # Read first 10 rows
        print("\n--- FIRST 15 ROWS ---")
        for r in range(1, min(16, max_row + 1)):
            row_vals = [sheet.cell(r, c).value for c in range(1, max_col + 1)]
            # Only print if not completely empty
            if any(val is not None for val in row_vals):
                print(f"Row {r:02d}: {row_vals}")
                
        # Read last 15 rows
        print("\n--- LAST 15 ROWS ---")
        start_row = max(1, max_row - 14)
        for r in range(start_row, max_row + 1):
            row_vals = [sheet.cell(r, c).value for c in range(1, max_col + 1)]
            if any(val is not None for val in row_vals):
                print(f"Row {r:02d}: {row_vals}")

        # Let's inspect some of the style formatting (e.g. font, fills, borders, alignments, column widths)
        print("\n--- COLUMNS INFO ---")
        for col_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            dim = sheet.column_dimensions[col_letter]
            print(f"Col {col_letter}: Width={dim.width}")

        # Inspect cell styling for typical cells (like row 1, row 5, row 8, final signature rows)
        print("\n--- SAMPLE CELLS STYLING ---")
        sample_cells = [(1, 1), (2, 1), (5, 1), (6, 1), (max_row - 2, 2) if max_row > 5 else (1, 1)]
        for r, c in sample_cells:
            cell = sheet.cell(r, c)
            font_str = f"Font(name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color.value if cell.font.color else None})"
            fill_str = f"Fill(type={cell.fill.fill_type}, color={cell.fill.start_color.value if cell.fill.start_color else None})"
            align_str = f"Align(horizontal={cell.alignment.horizontal}, vertical={cell.alignment.vertical})"
            print(f"Cell ({r},{c}) [{cell.value}]: {font_str} | {fill_str} | {align_str}")
