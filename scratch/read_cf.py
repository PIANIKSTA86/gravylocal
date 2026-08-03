import os
import openpyxl

ref_dir = r"c:\Users\JULIAN\Desktop\GravyLocal2.0\DatosReferencia\reportes"
filepath = os.path.join(ref_dir, "flujo de caja.xlsx")

wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb.active
print(f"Sheet: {sheet.title} (Rows: {sheet.max_row}, Cols: {sheet.max_column})")
for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
    if any(val is not None for val in row_vals):
        # Print non-empty rows
        non_empty = [f"Col{c}: {val}" for c, val in enumerate(row_vals, 1) if val is not None]
        print(f"Row {r:02d}: {row_vals[:10]}")
