import os
import openpyxl

ref_dir = r"c:\Users\JULIAN\Desktop\GravyLocal2.0\DatosReferencia\reportes"
files = [
    "Estado de resultados.xlsx",
    "estado de situacion financiera.xlsx"
]

for filename in files:
    filepath = os.path.join(ref_dir, filename)
    print(f"\n=========================================\nFILE: {filename}\n=========================================")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
        for r in range(1, min(45, sheet.max_row + 1)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(12, sheet.max_column + 1))]
            if any(val is not None for val in row_vals):
                print(f"Row {r:02d}: {row_vals}")
